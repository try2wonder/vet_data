import os
import json
import pdfplumber
import csv
import re
from openpyxl import Workbook, load_workbook


#Starting new idea with classses
class PetLab:
    def __init__(self, petname,labname):
        self.petname = petname
        self.labname = labname
    def showPetLab(self):
        return self.petname, self.labname
    def changePet(self, new_name):
        self.petname = new_name
    def changeLab(self, new_name):
        self.labname = new_name
    def addDataDay(self):
        return#this is the one that adds a column with data



#############Checking pets and labs. Adding new pets and labs
pets = {}

pet_name = ''
lab_name = []

wb = load_workbook('pet_lab_data.xlsx')

def fetchMetaData():
    json_file = 'petsAndLabs.json'
    global pets
    try:
        with open(json_file,'r') as file:
            content = file.read().strip()
            if content:
                pets = json.load(content)
                
            else:
                pets = {}
    except FileNotFoundError:
        print("JSON file not found, creating new one")
        pets = {}
    except JSONDecodeError:
        peint("Invalid JSON in file. Initializing empty pets dictionary")
        pets = {}




def pet_choice_dialog():
    global pet_name
    global lab_name
    pet_name = input("Enter pet's name: ")
    lab_name = input("Enter lab's name: ")
    

fetchMetaData()   

pet_choice_dialog()
print(pets, pet_name, lab_name)

def check_pet_lab(pet_name,lab_name):
    if pet_name in pets.keys():
        if lab_name in pets[pet_name]:
            return "pet lab found" #choosing the right Sheet
        else:
            choi=input("Pet is found, but not the lab. Do you want to add another pet_lab combination?(y/n): ")
            if choi == "y":
                pets[pet_name].append(lab_name)
                print(pets)
            else:
                print(pets)
                choi = input("Do you want to return to pet_lab choice?(y/n): ")
                if choi == "y":
                    pet_choice_dialog()
                    check_pet_lab(pet_name,lab_name)
    else:
        choi = input("no pet_lab combination found. Do you want to create a new one?(y/n): ")
        if choi == "y":
            pets[pet_name] = [lab_name]
        else:
            choi = input("Do you want to return to pet_lab choice?(y/n): ")
            if choi == "y":
                pet_choice_dialog()
                check_pet_lab(pet_name,lab_name)
                 
            print(pets)

check_pet_lab(pet_name, lab_name)


###########Creating MetaData
def writePetLabListChanges():
    with open('petsAndLabs.json','w')as file:
        json.dump(pets, file)

writePetLabListChanges()

#########Workign with pdf
pdf_name = "хорси.pdf"
def get_table(name):
    with pdfplumber.open(name) as pdf:
        page = pdf.pages[0]
    
        full_text = page.extract_text()
        tables = page.extract_tables()
        #print(type(text))   #string
        #print(type(tables)) #list
        return tables

def get_clean_dict(pdf_name):
    temp_list_var = get_table(pdf_name)[0]
    clean_list = temp_list_var[1:]
    clean_dict = {testd[0] : testd[2] for testd in clean_list}
    print(clean_dict)
    return clean_dict



##################Creating xlsx file with sections
def create_excel_with_sheets():
    wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    for pet, labs in pets.items():
        for lab in labs:
            sheet = wb.create_sheet(title=f"{pet}_{lab}")
            #Add headers and sample data and more rows
            sheet.append(["The date"])

    wb.save("pet_lab_data.xlsx")
    print("Created Excel file with multiple sheets")

choice1 = input("do you want to create a new .xlsx file(1), or use an existing one(2): ")
if int(choice1) == 1:
    create_excel_with_sheets()
elif int(choice1) == 2:
    print("checking existing one")

###Sheet work
def create_new_sheet(wb, pet_name, lab_name):
    ws1 = wb.create_sheet(title=f"{pet_name}_{lab_name}")
    ws1.append([f"{pet_name}_{lab_name} WorkSheet created"])
    wb.save('pet_lab_data.xlsx')

def get_or_create_sheet(wb, pet_name, lab_name):
    ###Gets existing sheet or creates new one if doesn't exist
    sheet_name = f"{pet_name}_{lab_name}"
    try:
        return wb[sheet_name]  # Try to get existing sheet
    except KeyError:
        print(f"Sheet '{sheet_name}' not found - creating new one")
        return create_new_sheet(wb, pet_name, lab_name)

########Making the first column (Table name, analisys names)
def make_first_column(wb, pet_name, lab_name):
    cd = get_clean_dict(pdf_name)
    sheet_name = f"{pet_name}_{lab_name})"
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(title=sheet_name)

    sheet.cell(row=1, column=1, value=f"{pet_name} {lab_name}")
    row_num = 2
    for analysis_name in cd.keys():
        sheet.cell(row=row_num, column=1, value=analysis_name)
        row_num += 1

make_first_column(wb, pet_name, lab_name)

########Adding Columns to Specific Sheets
def add_data_column(wb, pet_name,lab_name):

    sheet = wb[f"{pet_name}_{lab_name}"]

    last_col = sheet.max_column #Finds the last one with data. Need to check if the date already exists there

    new_columns = {
            last_col+1: "New Column 1" #This is supposed to be the date of analysis. mby make a temp var for column number, in case its not the last one
            }

    for col_num, header in new_columns.items():
        sheet.cell(row=1, column=col_num, value=header)
    
    cd = get_clean_dict(pdf_name)
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=last_col+1, value=f"Data {row-1}")#Data row 1 is the values i need to get from PDF

    wb.save('pet_lab_data.xlsx')

add_data_column(wb, pet_name,lab_name)


