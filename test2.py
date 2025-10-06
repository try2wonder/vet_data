from openpyxl import Workbook, load_workbook

pet_name = 'test'
lab_name = 'page'

wb = load_workbook('pet_lab_data.xlsx')

ws1 = wb.create_sheet(title=f"{pet_name}_{lab_name}")
ws1.append(["Test page"])

wb.save('pet_lab_data.xlsx')


